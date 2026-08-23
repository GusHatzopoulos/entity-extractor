from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from src.entity.final_selector import (
    select_locations,
    select_persons,
)
from src.entity.types import EntityRecord


def _sort_entities(
    entities: list[EntityRecord],
) -> list[EntityRecord]:
    """
    Sort entities by occurrence count descending,
    then alphabetically by canonical name.
    """

    return sorted(
        entities,
        key=lambda entity: (
            -entity.get("occurrences", 0),
            entity.get("entity", ""),
        ),
    )


def _write_appendix_sheet(
    worksheet: Worksheet,
    entities: list[EntityRecord],
) -> None:
    """
    Write a simple appendix sheet containing the canonical
    entity name and its occurrence count.
    """

    worksheet.append(
        [
            "Name",
            "Occurrences",
        ]
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for entity in entities:
        worksheet.append(
            [
                entity.get("entity", ""),
                entity.get("occurrences", 0),
            ]
        )

    worksheet.column_dimensions["A"].width = 40
    worksheet.column_dimensions["B"].width = 15

    worksheet.freeze_panes = "A2"

    if worksheet.max_row >= 1:
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


def export_appendix_to_xlsx(
    entities: list[EntityRecord],
    output_path: str | Path,
) -> None:
    """
    Export final canonical book entities into two sheets:

        Characters
        Locations

    The supplied entities are expected to have already
    passed final selection.
    """

    path = Path(output_path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    persons = _sort_entities(
        select_persons(entities)
    )

    locations = _sort_entities(
        select_locations(entities)
    )

    workbook = Workbook()

    characters_sheet = workbook.active

    if characters_sheet is None:
        raise RuntimeError(
            "Could not create the Characters worksheet."
        )

    characters_sheet.title = "Characters"

    locations_sheet = workbook.create_sheet(
        title="Locations"
    )

    _write_appendix_sheet(
        characters_sheet,
        persons,
    )

    _write_appendix_sheet(
        locations_sheet,
        locations,
    )

    workbook.save(path)