from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.entity.types import EntityRecord


HEADERS = [
    "entity",
    "original_type",
    "type",
    "source",
    "occurrences",
    "confidence",
    "person_score",
    "location_score",
    "classification_reason",
    "keep",
]


def _write_sheet(
    workbook: Workbook,
    title: str,
    entities: list[EntityRecord],
) -> None:
    worksheet = workbook.create_sheet(title=title)

    worksheet.append(HEADERS)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for entity in entities:
        worksheet.append(
            [
                entity.get("entity", ""),
                entity.get("original_type", ""),
                entity.get("type", ""),
                entity.get("source", ""),
                entity.get("occurrences", 0),
                entity.get("confidence", ""),
                entity.get("person_score", 0),
                entity.get("location_score", 0),
                entity.get("classification_reason", ""),
                entity.get("keep", ""),
            ]
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = 0

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        column_letter = get_column_letter(
            column_cells[0].column
        )

        worksheet.column_dimensions[
            column_letter
        ].width = min(max_length + 2, 60)


def export_entities_to_xlsx(
    entities: list[EntityRecord],
    output_path: str | Path,
) -> None:
    path = Path(output_path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    default_sheet = workbook.active

    if default_sheet is not None:
        workbook.remove(default_sheet)

    persons = [
        entity
        for entity in entities
        if entity.get("type") == "PERSON"
    ]

    locations = [
        entity
        for entity in entities
        if entity.get("type") in {"LOCATION", "GPE"}
    ]

    organizations = [
        entity
        for entity in entities
        if entity.get("type") == "ORG"
    ]

    review = [
        entity
        for entity in entities
        if (
            entity.get("type") == "UNKNOWN"
            or entity.get("confidence") in {"LOW", "REVIEW", ""}
        )
    ]

    _write_sheet(
        workbook,
        "All Entities",
        entities,
    )

    _write_sheet(
        workbook,
        "Persons",
        persons,
    )

    _write_sheet(
        workbook,
        "Locations",
        locations,
    )

    _write_sheet(
        workbook,
        "Organizations",
        organizations,
    )

    _write_sheet(
        workbook,
        "Review",
        review,
    )

    workbook.save(path)